# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from bson import json_util
import codecs
import datetime
import json
import re
from simpleText import replacePol
from getSAOS import getSOASjQuery

data = "klauzule_min_160923.xlsx"
out = "klauzulePython.json"

naglowki = ['lp','wyrokData','syg','sad','powod','pozwani','klauzula','wpisData','uwagi','branza']
dodatkowe = ['parent','saos','wyszukiwanie']

zleZnaki = {
    u'\'' : u'',
    u'\"' : u'',
    u'\u2013' :u'-',
    u'\xa7' : u'', # pownien byc znak paragrafu
    u'\u201d' : u'',
    u'\u201D' : u'',
    u'\u201e' : u'',
    u'\u2026' : u'',
}

def usunZleZnaki(text):
    for kk in zleZnaki.keys():
        text2 = text.replace(kk, zleZnaki[kk])
        text = text2
    return text

def przeprocesujKlauzule(string):
    return string[1:-1]

def sygnaturaAktDlaSAOS(syg):
    """zwracana krotka: <rzymskie> <kodSadu> <numer>/<rok>"""
    pattern = r'(\w+) (\w+)\s*(\d+)/(\d+)'
    m = re.findall(pattern, syg, re.LOCALE)
    saos = ''
    if m:
        krotka = m[0]
        saos = getSOASjQuery(krotka[0], krotka[1], krotka[2], krotka[3])
        #print saos
    else:
        print "WARNING!!!"
        print syg
        pass
    return saos


def wersjaDoWyszukiwania(tekst):
    tekst2 = tekst.lower()
    slowa = re.findall(r'(\w+)', tekst2, re.UNICODE)
    #print slowa
    out = u' '.join(slowa)
    
    #uzmiennienie ze wzgledu na liczby
    bezLiczb = re.sub(r'\d+', u'', out)
    #print bezLiczb
    return bezLiczb

def znajdzGwiazdki(tekst):
    print tekst.count(u'...')
    print tekst.count(u'...')

def grabData(sheet):
    data = []
    ii = 0
    for row in w.iter_rows(min_row=2, max_col=10, max_row=6627): # wszystkie rzedy poza naglowkami
        line = []
        skipRow = False
        for jj,cell in enumerate(row):
            value = cell.value

            if value == None and jj ==6: # ta linia mowi o tym, ze nie mamy naglowka
                skipRow = True
                break
            if not value:
                line.append(u'')
            elif jj == 0:
                line.append(str(value))
            elif (jj == 1 or jj == 7) and value:
                #print value
                if type(value) == type(u'test'):
                    line.append(value)
                else:
                    line.append(str(value.date()))
                #     line[naglowki[jj]] = str(value)
            elif jj == 6: # klauzule maja zwyczaj miec " na poczatku i na koncu
                line.append(usunZleZnaki(value))
                
                #print przeprocsujKlauzule(cell.value)
            elif jj == 0:
                line.append(value)
            else:
                line.append(value)
                
        if skipRow: # musimy wyskoczyc z dwoch petli
            continue

        line.append(line[0]) # dla kazdej bazowej klauzuli jego rodzic to numer porzadkowy na liscie
        # to sie przyda do synonimow
        #line['wyszukiwanie'] = wersjeDoWyszukiwania(line['klauzula']) # wersje dla wyszukiwarki
        line.append(sygnaturaAktDlaSAOS(line[2]))
        line.append(wersjaDoWyszukiwania(line[6]))
        #if not line['saos']:
        #    print line['lp']
        data.append(line)
        ii += 1
    return data


def genSaosHTML(link):
    return u'<a href="%s">Sprawdz SAOS</a>' % link

def writeOut(data, out):
    #calosc = u''
    calosc = u"""
    <table class = 'pure-table'>
        <tr>
            <th><img src="https://uokik.gov.pl/img/d1df7e93b44210b96599fc24d70c5bc0.gif" height=109/></th>
            <th><img src="https://www.saos.org.pl/static/image/saosLogo.png" height=109/></th>
        </tr>
    </table>
    <p>UWAGA system SAOS niestety w przypadku orzeczen klauzul nie dozwolonych nie jest kompletny!</p>
    <table class=\'pure-table\'> \n"""
    templateLin = u"""\t<tr %s >
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
        <th>%s</th>
    </tr>"""
    calosc = calosc + u"""\t<tr>
        <th>Lp</th>
        <th>Data wyroku</th>
        <th>Syg.</th>
        <th>SAOS</th>
        <th>Sad</th>
        <th>Powod</th>
        <th>Pozwany</th>
        <th>Postanowanie niedozwolone</th>
        <th>Data wpisu</th>
        <th>Uwagi</th>
        <th>Branza</th>
    </tr>\n"""
    ii = 0 
    data.reverse()
    for line in data:
        #calosc = u'
        classText = u''
        if ii % 2 == 0:
            classText = 'class=\'pure-table-odd\''
        lin = templateLin % (classText, unicode(line[0]), line[1], line[2], genSaosHTML(sygnaturaAktDlaSAOS(line[2])), line[3], line[4],line[5],line[6],line[7],line[8],line[9])
        calosc = calosc +lin
        ii = ii + 1
        #print u';'.join(l)
    #print data
    calosc = calosc + u'</table> \n'

    file = codecs.open("klauzuleLista.html", "w", "utf-8")
    file.write(calosc)
    file.close()
    
if __name__ == "__main__":
    wb = load_workbook(filename = 'klauzule_min_160923.xlsx')
    w = wb[wb.get_sheet_names()[0]]
    data = grabData(w)
    writeOut(data, out)
