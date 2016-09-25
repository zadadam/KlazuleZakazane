# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from bson import json_util
import codecs
import datetime
import json
import re
from simpleText import replacePol
from getSAOS import getSOASjQuery

data = "klauzule_min_160923.xlsx"
out = "klauzulePython.json"

naglowki = ['lp','wyrokData','syg','sad','powod','pozwani','klauzula','wpisData','uwagi','branza']
dodatkowe = ['parent','wyszukiwanie','saos']

zleZnaki = {
    u'\'' : u'',
    u'\"' : u'',
    u'\u2013' :u'-',
    u'\xa7' : u'', # pownien byc znak paragrafu
    u'\u201d' : u'',
    u'\u201D' : u'',
    u'\u201e' : u'',
    u'\u2026' : u'',
}

def usunZleZnaki(text):
    for kk in zleZnaki.keys():
        text2 = text.replace(kk, zleZnaki[kk])
        text = text2
    return text

def przeprocesujKlauzule(string):
    return string[1:-1]

def sygnaturaAktDlaSAOS(syg):
    """zwracana krotka: <rzymskie> <kodSadu> <numer>/<rok>"""
    pattern = r'(\w+) (\w+)\s*(\d+)/(\d+)'
    m = re.findall(pattern, syg, re.LOCALE)
    saos = ''
    if m:
        krotka = m[0]
        saos = getSOASjQuery(krotka[0], krotka[1], krotka[2], krotka[3])
        #print saos
    else:
        print "WARNING!!!"
        print syg
        pass
    return saos


def wersjaDoWyszukiwania(tekst):
    tekst2 = tekst.lower()
    slowa = re.findall(r'(\w+)', tekst2, re.UNICODE)
    #print slowa
    out = u' '.join(slowa)
    
    #uzmiennienie ze wzgledu na liczby
    bezLiczb = re.sub(r'\d+', u'', out)
    #print bezLiczb
    return bezLiczb

def znajdzGwiazdki(tekst):
    print tekst.count(u'...')
    print tekst.count(u'...')

def grabData(sheet):
    data = []
    ii = 0
    for row in w.iter_rows(min_row=2, max_col=10, max_row=6627): # wszystkie rzedy poza naglowkami
        line = []
        skipRow = False
        for jj,cell in enumerate(row):
            value = cell.value

            if value == None and jj ==6: # ta linia mowi o tym, ze nie mamy naglowka
                skipRow = True
                break
            if not value:
                line.append("")
            if jj == 0:
                line.append(str(value))
            elif (jj == 1 or jj == 7) and value:
                #print value
                if type(value) == type(u'test'):
                    line.append(value)
                else:
                    line.append(str(value.date()))
                #     line[naglowki[jj]] = str(value)
            elif jj == 6: # klauzule maja zwyczaj miec " na poczatku i na koncu
                line.append(usunZleZnaki(value))
                
                #print przeprocsujKlauzule(cell.value)
            elif jj == 0:
                line.append(value)
            else:
                line.append(value)
                
        if skipRow: # musimy wyskoczyc z dwoch petli
            continue

        line.append(line[0]) # dla kazdej bazowej klauzuli jego rodzic to numer porzadkowy na liscie
        # to sie przyda do synonimow
        #line['wyszukiwanie'] = wersjeDoWyszukiwania(line['klauzula']) # wersje dla wyszukiwarki
        line.append(sygnaturaAktDlaSAOS(line[2]))
        line.append(wersjaDoWyszukiwania(line[6]))
        #if not line['saos']:
        #    print line['lp']
        data.append(line)
        ii += 1
    return data





def writeOut(data, out):
    calosc = u''
    for line in data:
        linia = u''
        for ii,e in enumerate(line):
            if ii != 0:
                linia += u';'
            linia = linia + unicode(e)
        calosc = calosc +linia +'\n'
        #print u';'.join(l)
    #print data

    file = codecs.open("out.csv", "w", "utf-8")
    file.write(calosc)
    file.close()
    
if __name__ == "__main__":
    wb = load_workbook(filename = 'klauzule_min_160923.xlsx')
    w = wb[wb.get_sheet_names()[0]]
    data = grabData(w)
    writeOut(data, out)
